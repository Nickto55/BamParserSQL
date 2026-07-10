import xlrd
import openpyxl
import pandas as pd

from pathlib import Path
from tkinter import messagebox
from typing import Optional, Union, List, Dict, Any, Hashable


class ExcelReader:
    """
    Базовый класс для чтения Excel-файлов.
    Автоматически определяет формат (.xls / .xlsx) и использует правильный парсер.
    Для старых .xls файлов использует xlrd с поддержкой кодировок Windows-1251/CP866.
    """

    ENCODINGS = ["cp1251", "utf-8", "cp866", "koi8_r", "iso-8859-5"]

    def __init__(
            self
            , file_path: str
            , sheet_name: Optional[Union[str, int]] = None
            , color_filter_column: Optional[str] = None
            , track_sheet_origin: bool = False
            , encoding: Optional[str] = None
            , header_row: int = 0
            , skip_sheet: Optional[Union[str, int, List[Union[str, int]]]] = None
    ):
        self.file_path = file_path
        self.sheet_name = sheet_name
        self.encoding = encoding
        self.header_row = header_row
        self.data: Optional[pd.DataFrame] = None
        self.columns_save: List[str] = []
        self.filtered_data: Optional[Dict[int, Dict[str, Any]]] = None
        self.color_filter_column = color_filter_column
        self.track_sheet_origin = track_sheet_origin
        self.sheet_origin: Optional[str] = None
        self.skip_sheet = skip_sheet
        self.cell_colors: Dict[tuple, str] = {}

        self.load_excel()

    @staticmethod
    def fix_encoding(text: Any) -> Any:
        """
        Исправляет типичные проблемы с кодировкой (cp1252->cp1251).
        Если текст уже корректный — возвращает как есть.
        """
        if not isinstance(text, str):
            return text

        try:
            return text.encode('cp1252', errors='ignore').decode('cp1251', errors='ignore')
        except (UnicodeEncodeError, UnicodeDecodeError):
            return text

    @staticmethod
    def fix_dataframe_encoding(df: pd.DataFrame) -> pd.DataFrame:
        """
        Рекурсивно исправляет кодировку во всех строковых значениях DataFrame.
        """
        df_fixed = df.copy()
        for col in df_fixed.columns:
            if df_fixed[col].dtype == object:
                df_fixed[col] = df_fixed[col].apply(ExcelReader.fix_encoding)
        return df_fixed

    def _detect_format(self, file_path: str) -> str:
        """Определяет формат файла по расширению или бинарной сигнатуре."""
        path = Path(file_path)
        ext = path.suffix.lower()
        if ext == ".xls":
            return "xls"
        elif ext in (".xlsx", ".xlsm"):
            return "xlsx"
        else:
            with open(file_path, "rb") as f:
                header = f.read(8)
            if header[:8] == b'\xd0\xcf\x11\xE0\xA1\xB1\x1A\xE1':
                return "xls"
            elif header[:4] == b'PK\x03\x04':
                return "xlsx"
            else:
                raise ValueError(f"Неизвестный формат файла: {file_path}")

    def _should_skip_sheet(self, sheet_name: str, sheet_index: int) -> bool:
        """Проверяет, нужно ли пропустить лист по имени или индексу."""
        if self.skip_sheet is None:
            return False

        skip_list = self.skip_sheet if isinstance(self.skip_sheet, list) else [self.skip_sheet]

        for skip_item in skip_list:
            if isinstance(skip_item, int) and sheet_index == skip_item:
                return True
            if isinstance(skip_item, str) and sheet_name == skip_item:
                return True

        return False

    def _clamp_header_row(self, max_rows: int) -> int:
        """
        Ограничивает header_row, чтобы он не выходил за пределы файла.
        Если header_row >= max_rows — сбрасывает в 0.
        """
        if self.header_row >= max_rows:
            print(f"[ExcelReader] WARNING: header_row={self.header_row} превышает количество строк ({max_rows}). "
                  f"Сбрасываем в 0.")
            return 0
        return self.header_row

    def load_excel(
            self
            , file_path: Optional[str] = None
            , sheet_name: Optional[Union[str, int]] = None
            , color_filter_column: Optional[str] = None
            , encoding: Optional[str] = None
            , skip_sheet: Optional[Union[str, int, List[Union[str, int]]]] = None
    ) -> Optional[pd.DataFrame]:
        """
        Загружает данные из Excel файла с автоматическим определением формата
        и исправлением кодировки.
        """
        path = file_path or self.file_path
        sheet = sheet_name if sheet_name is not None else self.sheet_name
        color_col = color_filter_column or self.color_filter_column
        enc = encoding or self.encoding
        hdr = self.header_row

        if skip_sheet is not None:
            self.skip_sheet = skip_sheet

        file_format = self._detect_format(path)
        print(f"[excel_reader] Обнаружен формат: {file_format}")

        try:
            if file_format == "xls":
                data = self._load_xls(path, sheet, color_col, enc, hdr)
            else:
                data = self._load_xlsx(path, sheet)

            if isinstance(data, pd.DataFrame):
                data = self.fix_dataframe_encoding(data)
            elif isinstance(data, dict):
                for key in data:
                    if isinstance(data[key], pd.DataFrame):
                        data[key] = self.fix_dataframe_encoding(data[key])

                sheet_names = list(data.keys())
                if sheet_names:
                    first_sheet = sheet_names[0]
                    print(f"Лист не указан. Загружаем первый лист: {first_sheet}")
                    data = data[first_sheet]
                else:
                    raise ValueError("Файл Excel не содержит листов.")

            if file_path is None and sheet_name is None:
                self.data = data
                print(f"Лист успешно загружен: {sheet if sheet else 'первый лист'}")

            return data, self.cell_colors

        except Exception as e:
            print(f"Ошибка при загрузке файла или листа: {e}")
            if file_path is None and sheet_name is None:
                self.data = None
            return None

    def _load_xls(self, path: str, sheet: Optional[Union[str, int]],
                  color_col: Optional[str], encoding: Optional[str],
                  header_row: int = 0) -> pd.DataFrame:
        """
        Загружает бинарный .xls через xlrd с поддержкой кодировок.
        Интегрированная логика: автоопределение кодировки, обработка пустых строк как NaN,
        поддержка произвольной строки заголовков.
        """
        print("[ExcelReader] Чтение как .xls (xlrd)...")

        workbook = None
        used_encoding = None
        last_error = None

        if encoding:
            try:
                workbook = xlrd.open_workbook(path, encoding_override=encoding)
                used_encoding = encoding
                print(f"[ExcelReader] .xls открыт с кодировкой: {encoding}")
            except Exception as e:
                print(f"[ExcelReader] Кодировка {encoding} не подошла: {e}")
                last_error = e
                workbook = None

        if workbook is None:
            for enc in self.ENCODINGS:
                try:
                    workbook = xlrd.open_workbook(path, encoding_override=enc)
                    used_encoding = enc
                    print(f"[ExcelReader] .xls открыт с кодировкой: {enc}")
                    break
                except Exception as e:
                    last_error = e
                    continue


        if workbook is None:
            try:
                workbook = xlrd.open_workbook(path)
                used_encoding = "auto"
                print("[ExcelReader] .xls открыт без принудительной кодировки")
            except Exception as e:
                raise ValueError(f"Не удалось открыть .xls файл. Последняя ошибка: {last_error}")

        available_sheets = workbook.sheet_names()

        if sheet is not None:
            if isinstance(sheet, int):
                sheet_idx = sheet
                self.sheet_origin = available_sheets[sheet]
            else:
                sheet_idx = available_sheets.index(sheet)
                self.sheet_origin = sheet
        else:
            sheet_idx = None
            for idx, s_name in enumerate(available_sheets):
                if not self._should_skip_sheet(s_name, idx):
                    sheet_idx = idx
                    self.sheet_origin = s_name
                    print(f"[ExcelReader] Пропущены листы: {self.skip_sheet}. Загружен первый доступный: {s_name}")
                    break

            if sheet_idx is None:
                raise ValueError(f"Все листы исключены (skip_sheet={self.skip_sheet}). Нет доступных листов.")

        ws = workbook.sheet_by_index(sheet_idx)

        effective_header = self._clamp_header_row(ws.nrows)

        data = []
        for row_idx in range(ws.nrows):
            row_data = []
            for col_idx in range(ws.ncols):
                cell = ws.cell(row_idx, col_idx)
                val = cell.value
                if val == '':
                    val = float('nan')
                row_data.append(val)
            data.append(row_data)


        if effective_header is not None and effective_header < len(data):
            headers = data[effective_header]
            for i, h in enumerate(headers):
                if pd.isna(h) or h == '' or h is None or str(h).strip() == '':
                    headers[i] = f'Unnamed: {i}'
            df_data = data[effective_header + 1:]
            df = pd.DataFrame(df_data, columns=headers)
        else:
            df = pd.DataFrame(data)

        df = self.fix_dataframe_encoding(df)

        if self.track_sheet_origin:
            df['__sheet_origin__'] = self.sheet_origin

        return df

    def _load_xlsx(self, path: str, sheet: Optional[Union[str, int]]) -> pd.DataFrame:

        """Загружает .xlsx через openpyxl/pandas."""
        return self._load_xlsx_with_colors(path, sheet)


    def _load_xlsx_with_colors(self, path: str, sheet: Optional[Union[str, int]]) -> pd.DataFrame:
        """Загружает данные и собирает цвета всех ячеек."""

        try:
            wb_check = openpyxl.load_workbook(path, data_only=True)
            if sheet is None:
                ws_check = wb_check.active
            elif isinstance(sheet, int):
                ws_check = wb_check.worksheets[sheet]
            else:
                ws_check = wb_check[sheet]
            max_rows = ws_check.max_row
            wb_check.close()
        except Exception:
            max_rows = 999999

        effective_header = self._clamp_header_row(max_rows) if max_rows < 999999 else self.header_row

        try:
            raw_data = pd.read_excel(path, sheet_name=sheet, header=effective_header)
        except ValueError as e:
            if "Passed header=" in str(e) and "but only" in str(e):
                print(f"[ExcelReader] WARNING: {e}. Пробуем с header=0.")
                raw_data = pd.read_excel(path, sheet_name=sheet, header=0)
            else:
                raise

        if isinstance(raw_data, dict):
            if self.skip_sheet is not None:
                skip_list = self.skip_sheet if isinstance(self.skip_sheet, list) else [self.skip_sheet]
                filtered_sheets = {}
                for s_name, s_df in raw_data.items():
                    try:
                        wb_temp = openpyxl.load_workbook(path, data_only=True)
                        sheet_names = wb_temp.sheetnames
                        wb_temp.close()
                        s_idx = sheet_names.index(s_name)
                    except:
                        s_idx = -1

                    if s_name not in skip_list and s_idx not in skip_list:
                        filtered_sheets[s_name] = s_df
                    else:
                        print(f"[ExcelReader] Пропущен лист: {s_name}")

                raw_data = filtered_sheets

            sheet_names = list(raw_data.keys())
            if sheet_names:
                first_sheet = sheet_names[0]
                print(f"[ExcelReader] Лист не указан. Загружаем первый лист: {first_sheet}")
                df = raw_data[first_sheet]
                sheet_for_colors = first_sheet
            else:
                raise ValueError("Файл Excel не содержит листов (все исключены).")
        else:
            df = raw_data
            sheet_for_colors = sheet

        df = self.fix_dataframe_encoding(df)

        wb = openpyxl.load_workbook(path, data_only=True)

        if sheet_for_colors is None:
            ws = wb.active
            self.sheet_origin = ws.title
        elif isinstance(sheet_for_colors, int):
            ws = wb.worksheets[sheet_for_colors]
            self.sheet_origin = ws.title
        else:
            ws = wb[sheet_for_colors]
            self.sheet_origin = sheet_for_colors

        excel_header_row = effective_header + 1

        self.cell_colors = {}

        columns = list(df.columns)

        for row_idx in range(excel_header_row + 1, ws.max_row + 1):
            for col_idx in range(1, min(ws.max_column + 1, len(columns) + 1)):
                cell = ws.cell(row=row_idx, column=col_idx)

                color = None
                if cell.fill and cell.fill.fgColor:
                    rgb = cell.fill.fgColor.rgb
                    if rgb and rgb not in ('00000000', 'FFFFFFFF', None):
                        color = str(rgb)

                if color:
                    col_name = columns[col_idx - 1]
                    self.cell_colors[(row_idx, col_name)] = color

        wb.close()

        if self.track_sheet_origin:
            df['__sheet_origin__'] = self.sheet_origin

        return df

    def get_dict_all_data(self) -> Dict[int, Dict[str, Any]]:
        """Возвращает весь словарь данных из self.data."""
        if self.data is None:
            print("Данные не загружены.")
            return {}

        return {index: row.to_dict() for index, row in self.data.iterrows()}

    def get_headers(self) -> List[str]:
        """Возвращает список заголовков колонок."""
        if self.data is None:
            return []
        return list(self.data.columns)

    def filter_and_save_columns(self, columns_to_save: Union[str, List[str], tuple]):
        """Сохраняет значения из указанных столбцов в словарь."""
        if self.data is None:
            print("Данные не загружены.")
            return

        if isinstance(columns_to_save, str):
            columns_to_save = [columns_to_save]
        elif not isinstance(columns_to_save, list):
            columns_to_save = list(columns_to_save)

        self.columns_save = columns_to_save
        self.filtered_data = {}

        for index, row in self.data.iterrows():
            self.filtered_data[index] = {
                col: row[col] for col in columns_to_save if col in row
            }

    def get_filtered_data(self) -> Optional[Dict[int, Dict[str, Any]]]:
        """Возвращает отфильтрованные данные."""
        return self.filtered_data

    @staticmethod
    def is_empty(val: Any) -> bool:
        """Проверяет, является ли значение пустым."""
        return (val is None or
                (isinstance(val, str) and val.strip().lower() in ("", "nan", "n/a", "none", "-", "null")) or
                (hasattr(pd, 'isna') and pd.isna(val)))

    def get_column_values(self, get_column: str,
                          foc_mode: bool = False,
                          skip_condition: Optional[callable] = None) -> List[str]:
        """Извлекает уникальные значения из указанной колонки."""
        if get_column not in self.columns_save:
            messagebox.showerror(
                "Ошибка",
                f"Название столбцов не совпадают.\n{get_column} в {self.columns_save}"
            )
            return []

        result = []
        all_data_dict = self.get_dict_all_data() if foc_mode else None

        for key in self.filtered_data:
            if foc_mode and skip_condition is not None:
                row_data = all_data_dict.get(key, {})
                if skip_condition(row_data):
                    continue

            value = self.filtered_data[key].get(get_column)
            if value is None or self.is_empty(value):
                continue

            value_str = str(value).strip()
            if not value_str:
                continue

            items = value_str.replace(", ", "|").split("|")

            for item in items:
                item = item.strip()
                if not item:
                    continue

                if ":" in item and "-" in item:
                    item = item[:len(item) // 2 + 1]

                if item not in result:
                    result.append(item)

        result.sort(reverse=True)
        result.append("")
        return result

    def return_data(self) -> Optional[pd.DataFrame]:
        """Возвращает загруженные данные."""
        return self.data


class MultiSheetReader:
    """Класс для чтения нескольких листов из одного Excel-файла."""

    def __init__(self, file_path: str, skip_sheets: Optional[Union[str, int, List[Union[str, int]]]] = None):
        self.file_path = file_path
        self.sheets: Dict[str, Optional[pd.DataFrame]] = {}
        self.skip_sheets = skip_sheets

    def _should_skip(self, sheet_name: str, sheet_index: int) -> bool:
        """Проверяет, нужно ли пропустить лист."""
        if self.skip_sheets is None:
            return False

        skip_list = self.skip_sheets if isinstance(self.skip_sheets, list) else [self.skip_sheets]

        for skip_item in skip_list:
            if isinstance(skip_item, int) and sheet_index == skip_item:
                return True
            if isinstance(skip_item, str) and sheet_name == skip_item:
                return True

        return False

    def load_sheets(self, sheet_names: Optional[List[str]] = None) -> Dict[str, Optional[pd.DataFrame]]:
        """
        Загружает указанные листы из Excel-файла.
        Если sheet_names не указан — загружает все листы кроме skip_sheets.
        """
        try:
            xl = pd.ExcelFile(self.file_path)
            all_sheet_names = xl.sheet_names
        except Exception as e:
            print(f"Ошибка при открытии файла: {e}")
            return {}

        if sheet_names is None:
            sheet_names = []
            for idx, name in enumerate(all_sheet_names):
                if not self._should_skip(name, idx):
                    sheet_names.append(name)
                else:
                    print(f"[MultiSheetReader] Пропущен лист: {name}")
        else:
            filtered = []
            for idx, name in enumerate(all_sheet_names):
                if name in sheet_names and not self._should_skip(name, idx):
                    filtered.append(name)
                elif name in sheet_names and self._should_skip(name, idx):
                    print(f"[MultiSheetReader] Пропущен лист: {name}")
            sheet_names = filtered

        for sheet_name in sheet_names:
            try:
                data = pd.read_excel(self.file_path, sheet_name=sheet_name)

                data = ExcelReader.fix_dataframe_encoding(data)
                self.sheets[sheet_name] = data
            except Exception as e:
                print(f"Ошибка при загрузке листа {sheet_name}: {e}")
                self.sheets[sheet_name] = None

        return self.sheets

    def get_sheet_as_dict(self, sheet_name: str) -> dict[Hashable, dict] | None:
        """Возвращает данные листа в виде словаря."""
        data = self.sheets.get(sheet_name)
        if data is None:
            return None

        return {index: row.to_dict() for index, row in data.iterrows()}

    def get_sheet(self, sheet_name: str) -> Optional[pd.DataFrame]:
        """Возвращает DataFrame указанного листа."""
        return self.sheets.get(sheet_name)